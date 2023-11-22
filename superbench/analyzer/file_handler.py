# Copyright (c) Microsoft Corporation.
# Licensed under the MIT license.

"""A module for file related functions in analyzer."""

from pathlib import Path
import re
import json

import jsonlines
import pandas as pd
import yaml
from openpyxl.styles import Alignment
import markdown

from superbench.common.utils import logger


def read_raw_data(raw_data_path):
    """Read raw data from raw_data_path and store them in raw_data_df.

    Args:
        raw_data_path (str): the path of raw data jsonl file

    Returns:
        DataFrame: raw data, node as index, metric name as columns
    """
    p = Path(raw_data_path)
    raw_data_df = pd.DataFrame()
    if not p.is_file():
        logger.log_and_raise(
            exception=FileNotFoundError, msg='FileHandler: invalid raw data path - {}'.format(raw_data_path)
        )

    try:
        with p.open(encoding='utf-8') as f:
            for single_node_summary in jsonlines.Reader(f):
                raw_data_df = pd.concat([raw_data_df, pd.DataFrame([single_node_summary])], axis=0, ignore_index=True)
        raw_data_df = raw_data_df.rename(raw_data_df['node'])
        raw_data_df = raw_data_df.drop(columns=['node'])
    except Exception as e:
        logger.log_and_raise(exception=IOError, msg='Analyzer: invalid raw data format - {}'.format(str(e)))
    return raw_data_df


def read_rules(rule_file=None):
    """Read rule from rule yaml file.

    Args:
        rule_file (str, optional): The path of rule yaml file. Defaults to None.

    Returns:
        dict: dict object read from yaml file
    """
    default_rule_file = Path(__file__).parent / 'rule/default_rule.yaml'
    p = Path(rule_file) if rule_file else default_rule_file
    if not p.is_file():
        logger.log_and_raise(
            exception=FileNotFoundError, msg='FileHandler: invalid rule file path - {}'.format(str(p.resolve()))
        )
    baseline = None
    with p.open() as f:
        baseline = yaml.load(f, Loader=yaml.SafeLoader)
    return baseline


def read_baseline(baseline_file):
    """Read baseline from baseline json file.

    Args:
        baseline_file (str): The path of baseline json file.

    Returns:
        dict: dict object read from json file
    """
    p = Path(baseline_file)
    if not p.is_file():
        logger.log_and_raise(
            exception=FileNotFoundError, msg='FileHandler: invalid baseline file path - {}'.format(str(p.resolve()))
        )
    baseline = None
    with p.open() as f:
        baseline = json.load(f)
    return baseline


def output_excel_raw_data(writer, raw_data_df, sheet_name):
    """Output raw data into 'sheet_name' excel page.

    Args:
        writer (xlsxwriter): xlsxwriter handle
        raw_data_df (DataFrame): the DataFrame to output
        sheet_name (str): sheet name of the excel
    """
    # Output the raw data
    if isinstance(raw_data_df, pd.DataFrame) and not raw_data_df.empty:
        raw_data_df.to_excel(writer, sheet_name, index=True)
    else:
        logger.warning('FileHandler: excel_data_output - {} data_df is empty.'.format(sheet_name))


def output_excel_data_not_accept(writer, data_not_accept_df, rules):
    """Output data_not_accept_df into 'Not Accept' excel page.

    Args:
        writer (xlsxwriter): xlsxwriter handle
        data_not_accept_df (DataFrame): the DataFrame to output
        rules (dict): the rules of DataDiagnosis
    """
    # Get the xlsxwriter workbook objects and init the format
    workbook = writer.book
    color_format_red = workbook.add_format({'bg_color': '#FFC7CE', 'font_color': '#9C0006'})
    percent_format = workbook.add_format({'num_format': '0.00%'})

    # Output the not accept
    if isinstance(data_not_accept_df, pd.DataFrame):
        data_not_accept_df.to_excel(writer, 'Not Accept', index=True)
        if not data_not_accept_df.empty:
            row_start = 1
            row_end = max(row_start, len(data_not_accept_df))
            columns = list(data_not_accept_df.columns)
            worksheet = writer.sheets['Not Accept']

            for rule in rules:
                if 'function' in rules[rule]:
                    for metric in rules[rule]['metrics']:
                        # The column index of the metrics should start from 1
                        col_index = columns.index(metric) + 1
                        # Apply percent format for the columns whose rules are variance type.
                        if rules[rule]['function'] == 'variance':
                            worksheet.conditional_format(
                                row_start,
                                col_index,
                                row_end,
                                col_index,    # start_row, start_col, end_row, end_col
                                {
                                    'type': 'no_blanks',
                                    'format': percent_format
                                }
                            )
                        # Apply red format if the value violates the rule.
                        if rules[rule]['function'] == 'value' or rules[rule]['function'] == 'variance':
                            match = re.search(r'(>|<|<=|>=|==|!=)(.+)', rules[rule]['criteria'])
                            if not match:
                                continue
                            symbol = match.group(1)
                            condition = float(match.group(2))
                            worksheet.conditional_format(
                                row_start,
                                col_index,
                                row_end,
                                col_index,    # start_row, start_col, end_row, end_col
                                {
                                    'type': 'cell',
                                    'criteria': symbol,
                                    'value': condition,
                                    'format': color_format_red
                                }
                            )

        else:
            logger.warning('FileHandler: excel_data_output - data_not_accept_df is empty.')
    else:
        logger.log_and_raise(RuntimeError, msg='FileHandler: excel_data_output - data_not_accept_df is not DataFrame.')


def generate_md_table(data_df, header):
    """Generate table text in markdown format.

    | header[0] | header[1] |
    |     ----  | ----      |
    |     data  | data      |
    |     data  | data      |

    Args:
        data (DataFrame): the data in table
        header (list): the header of table

    Returns:
        list: lines of markdown table
    """
    lines = []
    data = data_df.values.tolist()
    max_width = len(max(data, key=len))
    header[len(header):max_width] = [' ' for i in range(max_width - len(header))]
    align = ['---' for i in range(max_width)]
    lines.append('| {} |\n'.format(' | '.join(header)))
    lines.append('| {} |\n'.format(' | '.join(align)))
    for line in data:
        full_line = [' ' for i in range(max_width)]
        full_line[0:len(line)] = [str(line[i]) for i in range(len(line))]
        lines.append('| {} |\n'.format(' | '.join(full_line)))
    return lines


def output_lines_in_md(lines, output_path):
    """Output lines in markdown format into a markdown file.

    Args:
        lines (list): lines in markdown format
        output_path (str): the path of output file
    """
    try:
        if len(lines) == 0:
            logger.warning('FileHandler: md_data_output is empty')
        with open(output_path, 'w') as f:
            f.writelines(lines)
    except Exception as e:
        logger.log_and_raise(exception=IOError, msg='FileHandler: md_data_output - {}'.format(str(e)))


def output_lines_in_html(lines, output_path):
    """Output markdown lines in html format file.

    Args:
        lines (list): lines in markdown format
        output_path (str): the path of output file
    """
    try:
        if len(lines) == 0:
            logger.warning('FileHandler: html_data_output is empty')
        lines = ''.join(lines)
        html_str = markdown.markdown(lines, extensions=['markdown.extensions.tables'])
        with open(output_path, 'w') as f:
            f.writelines(html_str)
    except Exception as e:
        logger.log_and_raise(exception=IOError, msg='FileHandler: html_data_output - {}'.format(str(e)))


def merge_column_in_excel(ws, row, column):
    """Merge cells in the selected index of column with continuous same contents.

    Args:
        ws (worksheet): the worksheet of the excel to process
        row (int): the max row index to merge
        column (int): the index of the column to merge
    """
    dict_from = {}
    aligncenter = Alignment(horizontal='center', vertical='center')
    # record continuous row index (start, end) with the same content
    for row_index in range(1, row + 1):
        value = str(ws.cell(row_index, column).value)
        if value not in dict_from:
            dict_from[value] = [row_index, row_index]
        else:
            dict_from[value][1] = dict_from[value][1] + 1
    # merge the cells
    for value in dict_from.values():
        if value[0] != value[1]:
            ws.merge_cells(start_row=value[0], start_column=column, end_row=value[1], end_column=column)
    # align center for merged cells
    for i in range(1, row + 1):
        ws.cell(row=i, column=column).alignment = aligncenter
